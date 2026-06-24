"""
fix_missing_page_urls.py

Applies known page URL corrections to a sport's _scp_master.xlsx.
The ODT-extracted URLs had slightly wrong slugs (missing panini- prefix,
'and' instead of '&', etc.). This script patches Column B with the
correct SCP page URLs so the scraper can retry those rows.

Run this, then re-run scrape_scp_other_sports.py for the same sport.

Usage:
    python scripts/fix_missing_page_urls.py football
    python scripts/fix_missing_page_urls.py hockey
    python scripts/fix_missing_page_urls.py basketball
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR  = REPO_ROOT / "data"
BASE      = "https://www.sportscardspro.com/console"

# ── URL correction rules per sport ────────────────────────────────────────────
#
# Each entry is (wrong_slug_fragment, correct_slug_fragment).
# Applied as a simple string replacement on Column B.
# Rules are checked in order — first match wins.
#
# {year} in comments is shorthand; the actual replacement operates on the
# full slug string so every year is covered automatically.

CORRECTIONS = {

    "football": [
        # Donruss Rookies & Highlights — must come BEFORE the plain donruss rule
        # because the plain slug is a substring of this one
        ("/football-cards-2025-donruss-rookies-and-highlights", "/football-cards-2025-panini-donruss-rookies-&-highlights"),

        # Donruss family — missing panini- prefix
        ("/football-cards-2018-donruss-elite",   "/football-cards-2018-panini-donruss-elite"),
        ("/football-cards-2019-donruss-elite",   "/football-cards-2019-panini-donruss-elite"),
        ("/football-cards-2020-donruss-elite",   "/football-cards-2020-panini-donruss-elite"),
        ("/football-cards-2021-donruss-elite",   "/football-cards-2021-panini-donruss-elite"),
        ("/football-cards-2022-donruss-elite",   "/football-cards-2022-panini-donruss-elite"),
        ("/football-cards-2023-donruss-elite",   "/football-cards-2023-panini-donruss-elite"),
        ("/football-cards-2024-donruss-elite",   "/football-cards-2024-panini-donruss-elite"),
        ("/football-cards-2025-donruss-elite",   "/football-cards-2025-panini-donruss-elite"),

        ("/football-cards-2020-donruss-optic",   "/football-cards-2020-panini-donruss-optic"),
        ("/football-cards-2021-donruss-optic",   "/football-cards-2021-panini-donruss-optic"),
        ("/football-cards-2022-donruss-optic",   "/football-cards-2022-panini-donruss-optic"),
        ("/football-cards-2023-donruss-optic",   "/football-cards-2023-panini-donruss-optic"),
        ("/football-cards-2024-donruss-optic",   "/football-cards-2024-panini-donruss-optic"),
        ("/football-cards-2025-donruss-optic",   "/football-cards-2025-panini-donruss-optic"),

        # Donruss plain — elite/optic must be fixed first (more specific) to avoid
        # this rule also matching those rows
        ("/football-cards-2020-donruss",         "/football-cards-2020-panini-donruss"),
        ("/football-cards-2021-donruss",         "/football-cards-2021-panini-donruss"),
        ("/football-cards-2022-donruss",         "/football-cards-2022-panini-donruss"),
        ("/football-cards-2023-donruss",         "/football-cards-2023-panini-donruss"),
        ("/football-cards-2024-donruss",         "/football-cards-2024-panini-donruss"),
        ("/football-cards-2025-donruss",         "/football-cards-2025-panini-donruss"),

        # Clearly Donruss — different set name on SCP
        ("/football-cards-2021-clearly-donruss", "/football-cards-2021-panini-chronicles-clearly-donruss-rated-rookies"),
        ("/football-cards-2022-clearly-donruss", "/football-cards-2022-panini-chronicles-clearly-donruss-rated-rookies"),

        # Score — missing panini- prefix
        ("/football-cards-2019-score",           "/football-cards-2019-panini-score"),
        ("/football-cards-2020-score",           "/football-cards-2020-panini-score"),
        ("/football-cards-2021-score",           "/football-cards-2021-panini-score"),
        ("/football-cards-2022-score",           "/football-cards-2022-panini-score"),
        ("/football-cards-2023-score",           "/football-cards-2023-panini-score"),
        ("/football-cards-2024-score",           "/football-cards-2024-panini-score"),
        ("/football-cards-2025-score",           "/football-cards-2025-panini-score"),

        # Rookies & Stars — 'and' → '&'
        ("/football-cards-2018-panini-rookies-and-stars", "/football-cards-2018-panini-rookies-&-stars"),
        ("/football-cards-2019-panini-rookies-and-stars", "/football-cards-2019-panini-rookies-&-stars"),
        ("/football-cards-2020-panini-rookies-and-stars", "/football-cards-2020-panini-rookies-&-stars"),
        ("/football-cards-2022-panini-rookies-and-stars", "/football-cards-2022-panini-rookies-&-stars"),
        ("/football-cards-2023-panini-rookies-and-stars", "/football-cards-2023-panini-rookies-&-stars"),
        ("/football-cards-2024-panini-rookies-and-stars", "/football-cards-2024-panini-rookies-&-stars"),
        ("/football-cards-2025-panini-rookies-and-stars", "/football-cards-2025-panini-rookies-&-stars"),

        # Zenith — missing panini- prefix
        ("/football-cards-2020-zenith",          "/football-cards-2020-panini-zenith"),
        ("/football-cards-2021-zenith",          "/football-cards-2021-panini-zenith"),
        ("/football-cards-2022-zenith",          "/football-cards-2022-panini-zenith"),

        # Finest — missing topps- prefix
        ("/football-cards-2024-finest",          "/football-cards-2024-topps-finest"),
        ("/football-cards-2025-finest",          "/football-cards-2025-topps-finest"),

        # National Treasures 2018 — had extra panini- prefix
        ("/football-cards-2018-panini-national-treasures", "/football-cards-2018-national-treasures"),

        # Plates & Patches — 'and' → '&'
        ("/football-cards-2018-panini-plates-and-patches", "/football-cards-2018-panini-plates-&-patches"),
        ("/football-cards-2019-panini-plates-and-patches", "/football-cards-2019-panini-plates-&-patches"),
        ("/football-cards-2020-panini-plates-and-patches", "/football-cards-2020-panini-plates-&-patches"),

    ],

    "hockey": [
        # Add hockey corrections here when ready
    ],

    "basketball": [
        # ── Haunted Hoops — must come before plain hoops rule (substring match) ──
        ("/basketball-cards-2023-hoops-haunted-hoops", "/basketball-cards-2023-panini-haunted-hoops"),
        ("/basketball-cards-2024-hoops-haunted-hoops", "/basketball-cards-2024-panini-haunted-hoops"),

        # ── Hoops Premium Stock — before plain hoops ──
        ("/basketball-cards-2019-hoops-premium-stock", "/basketball-cards-2019-panini-hoops-premium-stock"),
        ("/basketball-cards-2023-hoops-premium-stock", "/basketball-cards-2023-panini-hoops-premium-stock"),

        # ── Hoops Winter — before plain hoops ──
        ("/basketball-cards-2019-hoops-winter",        "/basketball-cards-2019-panini-hoops-winter"),
        ("/basketball-cards-2020-hoops-winter",        "/basketball-cards-2020-panini-hoops-winter"),
        ("/basketball-cards-2021-hoops-winter",        "/basketball-cards-2021-panini-hoops-winter"),
        ("/basketball-cards-2022-hoops-winter",        "/basketball-cards-2022-panini-hoops-winter"),
        ("/basketball-cards-2023-hoops-winter",        "/basketball-cards-2023-panini-hoops-winter"),
        ("/basketball-cards-2024-hoops-winter",        "/basketball-cards-2024-panini-hoops-winter"),

        # ── Hoops plain ──
        ("/basketball-cards-2019-hoops",               "/basketball-cards-2019-panini-hoops"),
        ("/basketball-cards-2020-hoops",               "/basketball-cards-2020-panini-hoops"),
        ("/basketball-cards-2021-hoops",               "/basketball-cards-2021-panini-hoops"),
        ("/basketball-cards-2022-hoops",               "/basketball-cards-2022-panini-hoops"),
        ("/basketball-cards-2023-hoops",               "/basketball-cards-2023-panini-hoops"),
        ("/basketball-cards-2024-hoops",               "/basketball-cards-2024-panini-hoops"),

        # ── Donruss family — missing panini- prefix ──
        # Donruss Elite before plain Donruss (substring match guard)
        ("/basketball-cards-2019-donruss-elite",       "/basketball-cards-2019-panini-donruss-elite"),
        ("/basketball-cards-2021-donruss-elite",       "/basketball-cards-2021-panini-donruss-elite"),
        ("/basketball-cards-2022-donruss-elite",       "/basketball-cards-2022-panini-donruss-elite"),

        # Donruss Optic before plain Donruss
        ("/basketball-cards-2019-donruss-optic",       "/basketball-cards-2019-panini-donruss-optic"),
        ("/basketball-cards-2020-donruss-optic",       "/basketball-cards-2020-panini-donruss-optic"),
        ("/basketball-cards-2021-donruss-optic",       "/basketball-cards-2021-panini-donruss-optic"),
        ("/basketball-cards-2022-donruss-optic",       "/basketball-cards-2022-panini-donruss-optic"),
        ("/basketball-cards-2023-donruss-optic",       "/basketball-cards-2023-panini-donruss-optic"),
        ("/basketball-cards-2024-donruss-optic",       "/basketball-cards-2024-panini-donruss-optic"),

        # Donruss plain
        ("/basketball-cards-2019-donruss",             "/basketball-cards-2019-panini-donruss"),
        ("/basketball-cards-2021-donruss",             "/basketball-cards-2021-panini-donruss"),
        ("/basketball-cards-2022-donruss",             "/basketball-cards-2022-panini-donruss"),
        ("/basketball-cards-2023-donruss",             "/basketball-cards-2023-panini-donruss"),
        ("/basketball-cards-2024-donruss",             "/basketball-cards-2024-panini-donruss"),

        # Clearly Donruss — user didn't confirm basketball slug, trying panini- prefix
        ("/basketball-cards-2019-clearly-donruss",     "/basketball-cards-2019-panini-clearly-donruss"),
        ("/basketball-cards-2020-clearly-donruss",     "/basketball-cards-2020-panini-clearly-donruss"),

        # ── Finest — missing topps- prefix ──
        ("/basketball-cards-2023-finest",              "/basketball-cards-2023-topps-finest"),
        ("/basketball-cards-2024-finest",              "/basketball-cards-2024-topps-finest"),

        # ── Panini Mosaic Prizm — word order reversed on SCP ──
        ("/basketball-cards-2018-panini-mosaic-prizm", "/basketball-cards-2018-panini-prizm-mosaic"),

        # ── Panini Top Class NBA — word order reversed on SCP ──
        ("/basketball-cards-2023-panini-top-class-nba", "/basketball-cards-2023-panini-nba-top-class"),

        # ── Topps Chrome Cosmic — word order reversed on SCP ──
        ("/basketball-cards-2023-topps-chrome-cosmic", "/basketball-cards-2023-topps-cosmic-chrome"),

        # ── Topps Chrome Sapphire — "edition" suffix not used on SCP ──
        ("/basketball-cards-2023-topps-chrome-sapphire-edition", "/basketball-cards-2023-topps-chrome-sapphire"),
        ("/basketball-cards-2024-topps-chrome-sapphire-edition", "/basketball-cards-2024-topps-chrome-sapphire"),
    ],
}


def apply_corrections(sport: str) -> None:
    xlsx_path = DATA_DIR / f"{sport}_scp_master.xlsx"
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found. Run the scraper first to create it.")
        sys.exit(1)

    rules = CORRECTIONS.get(sport, [])
    if not rules:
        print(f"No corrections defined for '{sport}' yet.")
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active

    changed = 0
    skipped = 0  # rows that already have a download URL — don't touch

    for row in range(2, ws.max_row + 1):
        page_url     = ws.cell(row=row, column=2).value
        download_url = ws.cell(row=row, column=3).value

        if not page_url:
            continue
        # Don't touch rows that already have a download URL
        if download_url:
            skipped += 1
            continue

        page_url = str(page_url).strip()
        original = page_url

        for wrong, correct in rules:
            if wrong in page_url:
                page_url = page_url.replace(wrong, correct)
                break  # first match wins

        if page_url != original:
            ws.cell(row=row, column=2).value = page_url
            name = ws.cell(row=row, column=1).value or f"Row {row}"
            print(f"  FIXED  {name}")
            print(f"    was: {original}")
            print(f"    now: {page_url}")
            changed += 1

    wb.save(xlsx_path)
    print(f"\n{changed} URL(s) corrected, {skipped} row(s) skipped (already have download URL).")
    if changed:
        print(f"Now rerun:  python scripts/scrape_scp_other_sports.py {sport}")


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python scripts/fix_missing_page_urls.py <sport>")
        sys.exit(1)
    apply_corrections(sys.argv[1].lower().strip())


if __name__ == "__main__":
    main()
