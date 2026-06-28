"""
scrape_inserts_download_urls.py

Reads 2023_Topps_Chrome_Inserts.xlsx, visits each SportsCardsPro page URL in
Column B that is missing a Download URL in Column C, grabs the
"Download Price List" href, and writes it back into Column C.

Saves after every single row so a crash loses at most one row of work.

Uses the same persistent Playwright browser context as the other SCP scrapers
(~/Library/Application Support/Google/Chrome/PlaywrightProfile). If you've
already logged in during a previous scrape session the login is reused
automatically.

Usage:
    python scripts/scrape_inserts_download_urls.py
"""

import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

XLSX_PATH = Path(
    "/Users/zachseabolt/Developer/Diamond in the rough Documents"
    "/Baseball/Baseball Seed Spreadsheet"
    "/2023 topps chrome baseball/2023_Topps_Chrome_Inserts.xlsx"
)

CHROME_USER_DATA_DIR = (
    Path.home()
    / "Library"
    / "Application Support"
    / "Google"
    / "Chrome"
    / "PlaywrightProfile"
)

COL_PAGE_URL = 2
COL_DOWNLOAD_URL = 3

DOWNLOAD_LINK_TEXT = "Download Price List"
PAGE_LOAD_TIMEOUT = 20_000
REQUEST_DELAY = 1.5


def scrape_download_url(page, scp_page_url: str) -> str | None:
    page.goto(scp_page_url, timeout=PAGE_LOAD_TIMEOUT, wait_until="domcontentloaded")
    link = page.get_by_role("link", name=DOWNLOAD_LINK_TEXT)
    href = link.first.get_attribute("href", timeout=5_000)
    return href or None


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {XLSX_PATH}")
        sys.exit(1)

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    rows_to_process = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        page_url = ws.cell(row=row_idx, column=COL_PAGE_URL).value
        download_url = ws.cell(row=row_idx, column=COL_DOWNLOAD_URL).value
        if page_url and not download_url:
            set_name = ws.cell(row=row_idx, column=1).value or f"Row {row_idx}"
            rows_to_process.append((row_idx, str(page_url).strip(), set_name))

    total = len(rows_to_process)
    if total == 0:
        print("All rows already have download URLs — nothing to do.")
        return

    print(f"{total} rows to process.")

    total_done = 0
    total_failed = 0

    with sync_playwright() as p:
        browser_context = p.chromium.launch_persistent_context(
            user_data_dir=str(CHROME_USER_DATA_DIR),
            channel="chrome",
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )

        page = browser_context.new_page()

        input(
            "Log into SportsCardsPro in the browser window if needed, "
            "then press Enter here to start scraping..."
        )

        for position, (row_idx, page_url, set_name) in enumerate(rows_to_process, start=1):
            prefix = f"{position}/{total} — {set_name}"

            try:
                href = scrape_download_url(page, page_url)

                if href:
                    ws.cell(row=row_idx, column=COL_DOWNLOAD_URL).value = href
                    wb.save(XLSX_PATH)
                    print(f"{prefix} — DONE ({href})")
                    total_done += 1
                else:
                    print(f"{prefix} — FAILED (no download link found on page)")
                    total_failed += 1

            except PlaywrightTimeout:
                print(f"{prefix} — FAILED (page timed out after {PAGE_LOAD_TIMEOUT / 1000}s)")
                total_failed += 1

            except Exception as exc:
                print(f"{prefix} — FAILED ({type(exc).__name__}: {exc})")
                total_failed += 1

            time.sleep(REQUEST_DELAY)

        browser_context.close()

    print("\n" + "─" * 60)
    print("SUMMARY")
    print(f"  Done (URL written):  {total_done}")
    print(f"  Failed (no link):    {total_failed}")
    print("─" * 60)

    if total_failed > 0:
        print(
            f"\n{total_failed} row(s) failed. Rerun the script to retry — "
            "rows with Column C still empty are picked up automatically."
        )


if __name__ == "__main__":
    main()
