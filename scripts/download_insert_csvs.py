"""
download_insert_csvs.py

Reads 2023_Topps_Chrome_Inserts.xlsx, downloads the CSV at each URL in
Column C, and saves it to:

    ~/Documents/Ripper Data/2023 Topps Chrome Inserts/

One file per insert set, named after the set (Column A).
Rate-limited to 1 download every 5 minutes (300 seconds) to respect the
SportsCardsPro limit.

Uses the persistent Playwright browser session (PlaywrightProfile) so the
SCP login cookie is reused automatically — no credentials needed.

Already-downloaded files are skipped automatically on reruns — safe to
interrupt and resume.

Usage:
    python scripts/download_insert_csvs.py
"""

import re
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ── Paths ──────────────────────────────────────────────────────────────────────

XLSX_PATH = Path(
    "/Users/zachseabolt/Developer/Diamond in the rough Documents"
    "/Baseball/Baseball Seed Spreadsheet"
    "/2023 topps chrome baseball/2023_Topps_Chrome_Inserts.xlsx"
)

OUTPUT_DIR = Path.home() / "Documents" / "Ripper Data" / "2023 Topps Chrome Inserts"

CHROME_USER_DATA_DIR = (
    Path.home()
    / "Library"
    / "Application Support"
    / "Google"
    / "Chrome"
    / "PlaywrightProfile"
)

# ── Settings ───────────────────────────────────────────────────────────────────

RATE_LIMIT_SECONDS = 300   # 5 minutes between downloads
PAGE_LOAD_TIMEOUT = 30_000

COL_SET_NAME = 1
COL_DOWNLOAD_URL = 3


def slugify(name: str) -> str:
    """Convert a set name into a safe filename (no special chars, spaces → underscores)."""
    name = name.strip()
    name = re.sub(r"[^\w\s-]", "", name)
    name = re.sub(r"\s+", "_", name)
    return name


def download_csv(page, download_url: str, output_path: Path) -> bool:
    """
    Navigate to the download URL and capture the file via Playwright's
    download event. page.goto throws when the server immediately returns a
    file attachment (no page to commit to) — that error is expected and safe
    to ignore. The download event fires regardless.
    """
    with page.expect_download(timeout=PAGE_LOAD_TIMEOUT) as download_info:
        try:
            page.goto(download_url, timeout=PAGE_LOAD_TIMEOUT, wait_until="commit")
        except Exception:
            # goto throws when the response is a file download, not a page.
            # The download event still fires — just let expect_download capture it.
            pass

    download = download_info.value
    download.save_as(str(output_path))
    return True


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {XLSX_PATH}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {OUTPUT_DIR}\n")

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        set_name = ws.cell(row=row_idx, column=COL_SET_NAME).value
        download_url = ws.cell(row=row_idx, column=COL_DOWNLOAD_URL).value
        if set_name and download_url:
            rows.append((str(set_name).strip(), str(download_url).strip()))

    total = len(rows)
    print(f"Found {total} rows with download URLs.")

    already_done = {f.stem for f in OUTPUT_DIR.glob("*.csv")}
    to_download = [(name, url) for name, url in rows if slugify(name) not in already_done]
    skipped = total - len(to_download)

    if skipped:
        print(f"Skipping {skipped} already-downloaded file(s).")

    if not to_download:
        print("All files already downloaded — nothing to do.")
        return

    remaining = len(to_download)
    total_minutes = (remaining - 1) * 5
    print(f"{remaining} file(s) to download. Estimated time: ~{total_minutes} minutes.\n")

    done = 0
    failed = 0

    with sync_playwright() as p:
        browser_context = p.chromium.launch_persistent_context(
            user_data_dir=str(CHROME_USER_DATA_DIR),
            channel="chrome",
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
            accept_downloads=True,
        )

        page = browser_context.new_page()

        input(
            "Log into SportsCardsPro in the browser window if needed, "
            "then press Enter here to start downloading..."
        )

        for position, (set_name, download_url) in enumerate(to_download, start=1):
            filename = slugify(set_name) + ".csv"
            output_path = OUTPUT_DIR / filename
            prefix = f"{position}/{remaining} — {set_name}"

            if position > 1:
                print(f"  Waiting {RATE_LIMIT_SECONDS // 60} minutes before next download...")
                time.sleep(RATE_LIMIT_SECONDS)

            print(f"{prefix}")
            print(f"  Downloading: {download_url}")

            try:
                download_csv(page, download_url, output_path)
                size_kb = output_path.stat().st_size / 1024
                print(f"  Saved: {filename} ({size_kb:.1f} KB)")
                done += 1

            except PlaywrightTimeout:
                print(f"  FAILED: timed out waiting for download")
                failed += 1
            except Exception as exc:
                print(f"  FAILED: {type(exc).__name__}: {exc}")
                failed += 1

            print()

        browser_context.close()

    print("─" * 60)
    print("SUMMARY")
    print(f"  Downloaded:  {done}")
    print(f"  Failed:      {failed}")
    print(f"  Skipped:     {skipped}")
    print("─" * 60)

    if failed:
        print(f"\n{failed} file(s) failed. Rerun to retry — successful downloads are skipped automatically.")


if __name__ == "__main__":
    main()
