"""
scrape_scp_other_sports.py

Reads SportsCardsPro page URLs from a per-sport .txt file (one URL per line),
builds a tracking spreadsheet, visits each page, grabs the "Download Price List"
href, and writes it back into the spreadsheet.

Works for football, hockey, basketball — any sport with a
data/scp-url-lists/{sport}-raw-urls.txt file from the extraction step.

FIRST RUN for a sport:
    Builds data/{sport}_scp_master.xlsx from the .txt file, then scrapes.

SUBSEQUENT RUNS:
    Reads the existing .xlsx, skips rows that already have a Download URL,
    and retries only the empty ones.

MISSING URLS:
    After the run, open the .xlsx. Rows where Column C is empty are missing
    their download link on SCP (either the page doesn't exist, isn't a set
    page, or the link text differs). Fill in the correct download URL manually
    for those rows and rerun — the scraper skips any row that already has a URL.

Usage:
    python scripts/scrape_scp_other_sports.py football
    python scripts/scrape_scp_other_sports.py hockey
    python scripts/scrape_scp_other_sports.py basketball

Requires:
    pip install playwright openpyxl
    playwright install chromium
"""

import re
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ── Paths ──────────────────────────────────────────────────────────────────────

REPO_ROOT = Path(__file__).resolve().parent.parent
URL_LIST_DIR = REPO_ROOT / "data" / "scp-url-lists"
DATA_DIR = REPO_ROOT / "data"

# Dedicated Playwright profile directory — same one the baseball scraper uses
# so you only need to log in once across all scraper runs.
CHROME_USER_DATA_DIR = (
    Path.home()
    / "Library"
    / "Application Support"
    / "Google"
    / "Chrome"
    / "PlaywrightProfile"
)

# ── Column indices (openpyxl is 1-based) ──────────────────────────────────────
COL_SET_NAME    = 1  # A — human-readable label derived from URL slug
COL_PAGE_URL    = 2  # B — SportsCardsPro set page
COL_DOWNLOAD_URL = 3  # C — Download Price List href (what we're filling in)

# ── Scraper settings ───────────────────────────────────────────────────────────
DOWNLOAD_LINK_TEXT = "Download Price List"
PAGE_LOAD_TIMEOUT  = 20_000   # ms
REQUEST_DELAY      = 1.5      # seconds between requests


# ── Helpers ────────────────────────────────────────────────────────────────────

def set_name_from_url(url: str, sport: str) -> str:
    """
    Derive a human-readable set name from an SCP page URL.

    Example:
        https://www.sportscardspro.com/console/football-cards-2018-panini-prizm
        → "2018 Panini Prizm"
    """
    path = url.rstrip("/").split("/")[-1]           # football-cards-2018-panini-prizm
    prefix = f"{sport}-cards-"
    if path.startswith(prefix):
        path = path[len(prefix):]                   # 2018-panini-prizm
    # Replace hyphens with spaces, title-case each word
    return " ".join(w.capitalize() for w in path.split("-"))


def build_xlsx(xlsx_path: Path, url_list_path: Path, sport: str) -> None:
    """Create a fresh tracking spreadsheet from a .txt URL list."""
    urls = [
        line.strip()
        for line in url_list_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = sport.capitalize()

    # Header row
    headers = ["Set Name", "SCP Page URL", "Download URL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # One row per URL
    for url in urls:
        name = set_name_from_url(url, sport)
        ws.append([name, url, None])

    # Reasonable column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 90

    wb.save(xlsx_path)
    print(f"Created {xlsx_path.name} with {len(urls)} rows.")


def scrape_download_url(page, scp_page_url: str) -> str | None:
    """Return the Download Price List href for this SCP page, or None."""
    page.goto(scp_page_url, timeout=PAGE_LOAD_TIMEOUT, wait_until="domcontentloaded")
    link = page.get_by_role("link", name=DOWNLOAD_LINK_TEXT)
    href = link.first.get_attribute("href", timeout=5_000)
    return href or None


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python scripts/scrape_scp_other_sports.py <sport>")
        print("  e.g. python scripts/scrape_scp_other_sports.py football")
        sys.exit(1)

    sport = sys.argv[1].lower().strip()
    url_list_path = URL_LIST_DIR / f"{sport}-raw-urls.txt"
    xlsx_path     = DATA_DIR / f"{sport}_scp_master.xlsx"

    if not url_list_path.exists():
        print(f"ERROR: URL list not found: {url_list_path}")
        print(f"Run extract_urls_from_odt.py first to generate {sport}-raw-urls.txt")
        sys.exit(1)

    # First run: build the spreadsheet from the txt file.
    # Subsequent runs: load the existing one and only retry empty rows.
    if not xlsx_path.exists():
        build_xlsx(xlsx_path, url_list_path, sport)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Collect rows that still need a download URL
    rows_to_process = []
    for row_idx in range(2, ws.max_row + 1):
        page_url     = ws.cell(row=row_idx, column=COL_PAGE_URL).value
        download_url = ws.cell(row=row_idx, column=COL_DOWNLOAD_URL).value
        set_name     = ws.cell(row=row_idx, column=COL_SET_NAME).value or f"Row {row_idx}"

        if page_url and not download_url:
            rows_to_process.append((row_idx, str(page_url).strip(), str(set_name)))

    total = len(rows_to_process)
    if total == 0:
        print(f"\n[{sport}] All rows already have download URLs — nothing to do.")
        print("Open the .xlsx to see which rows are still missing (Column C empty).")
        return

    print(f"\n[{sport}] {total} rows to scrape.")
    print(f"Spreadsheet: {xlsx_path}")
    print()

    total_done   = 0
    total_failed = 0

    with sync_playwright() as p:
        browser_context = p.chromium.launch_persistent_context(
            user_data_dir=str(CHROME_USER_DATA_DIR),
            channel="chrome",
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )

        page = browser_context.new_page()

        input(
            "Log into SportsCardsPro in the browser window if needed,\n"
            "then press Enter here to start scraping..."
        )

        for position, (row_idx, page_url, set_name) in enumerate(rows_to_process, start=1):
            prefix = f"[{sport}] {position}/{total} — {set_name}"

            try:
                href = scrape_download_url(page, page_url)

                if href:
                    ws.cell(row=row_idx, column=COL_DOWNLOAD_URL).value = href
                    wb.save(xlsx_path)
                    print(f"{prefix} — DONE")
                    total_done += 1
                else:
                    print(f"{prefix} — MISSING (no download link on page)")
                    total_failed += 1

            except PlaywrightTimeout:
                print(f"{prefix} — TIMEOUT ({PAGE_LOAD_TIMEOUT / 1000}s)")
                total_failed += 1

            except Exception as exc:
                print(f"{prefix} — ERROR ({type(exc).__name__}: {exc})")
                total_failed += 1

            time.sleep(REQUEST_DELAY)

        browser_context.close()

    # ── Summary ────────────────────────────────────────────────────────────────
    print("\n" + "─" * 60)
    print(f"SUMMARY — {sport}")
    print(f"  Download URL found : {total_done}")
    print(f"  Missing / failed   : {total_failed}")
    print(f"  Already had URL    : {ws.max_row - 1 - total}")
    print("─" * 60)

    if total_failed > 0:
        print(f"\n{total_failed} row(s) are missing a download URL.")
        print("Open the spreadsheet and look at Column C — empty rows are the missing ones.")
        print("To fix them:")
        print("  1. Find the correct SCP download URL manually (see notes below).")
        print("  2. Paste it into Column C for that row.")
        print("  3. Rerun this script — it skips rows that already have a URL.")
        print()
        print("Pattern for manual URLs:")
        print("  https://www.sportscardspro.com/game/[sport]-cards-[year]-[set-slug]?q=Download+Price+List")
        print("  (The exact path varies — check the SCP page for the correct href.)")


if __name__ == "__main__":
    main()
