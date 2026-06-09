"""
scrape_scp_download_urls.py

Phase 2 of the SCP data pipeline. Reads baseball_scp_master.xlsx, visits each
SportsCardsPro page URL in Column B that is missing a Download URL in Column C,
grabs the "Download Price List" href, and writes it back into Column C.

Saves after every single row so a crash loses at most one row of work.

Uses a persistent Playwright browser context with a dedicated profile at
~/Library/Application Support/Google/Chrome/PlaywrightProfile. On first run
you'll be prompted to log into SportsCardsPro manually; subsequent runs reuse
the saved session automatically. No credentials are hardcoded or read from this
script.

Usage:
    pip install playwright openpyxl
    playwright install chromium
    python scripts/scrape_scp_download_urls.py
"""

import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ── Paths ──────────────────────────────────────────────────────────────────────

XLSX_PATH = Path(__file__).parent.parent / "data" / "baseball_scp_master.xlsx"

# Dedicated Playwright profile directory. Isolated from your main Chrome profile
# so there's no conflict when Chrome is open. On first run, log in manually;
# the session is saved here for all future runs.
CHROME_USER_DATA_DIR = Path.home() / "Library" / "Application Support" / "Google" / "Chrome" / "PlaywrightProfile"

# ── Column indices (openpyxl uses 1-based columns) ────────────────────────────
# Column A = 1 (set name / label)
# Column B = 2 (SCP page URL)
# Column C = 3 (Download URL — this is what we're filling in)

COL_PAGE_URL = 2
COL_DOWNLOAD_URL = 3

# ── Selector for the Download Price List link ─────────────────────────────────
# SportsCardsPro renders a plain <a> tag whose visible text contains
# "Download Price List". We grab the href directly without clicking.
DOWNLOAD_LINK_TEXT = "Download Price List"

# How long to wait for the page to load (milliseconds)
PAGE_LOAD_TIMEOUT = 20_000

# Polite delay between requests so we don't hammer the server (seconds)
REQUEST_DELAY = 1.5


def scrape_download_url(page, scp_page_url: str) -> str | None:
    """
    Navigate to scp_page_url and return the href of the "Download Price List"
    link, or None if the link isn't found.
    """
    page.goto(scp_page_url, timeout=PAGE_LOAD_TIMEOUT, wait_until="domcontentloaded")

    # Look for an <a> whose text matches exactly or contains the target phrase.
    # Using get_by_role so we don't depend on brittle CSS class names.
    link = page.get_by_role("link", name=DOWNLOAD_LINK_TEXT)

    # If multiple matches, first() is fine — there's only ever one download link.
    href = link.first.get_attribute("href", timeout=5_000)
    return href or None


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {XLSX_PATH}")
        print("Place baseball_scp_master.xlsx in the data/ directory and rerun.")
        sys.exit(1)

    wb = load_workbook(XLSX_PATH)
    sheet_names = wb.sheetnames

    total_done = 0
    total_skipped = 0
    total_failed = 0

    with sync_playwright() as p:
        # launch_persistent_context keeps cookies/session from your real Chrome.
        # channel="chrome" uses the installed Chrome binary (not Chromium).
        # headless=False lets you see what's happening and avoids bot-detection
        # quirks that some sites apply to headless browsers.
        browser_context = p.chromium.launch_persistent_context(
            user_data_dir=str(CHROME_USER_DATA_DIR),
            channel="chrome",
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )

        page = browser_context.new_page()

        # First run: PlaywrightProfile won't have a session yet. Give the user
        # time to log in before scraping starts. Subsequent runs skip this
        # automatically because the session cookie is already saved in the profile.
        input("Log into SportsCardsPro in the browser window, then press Enter here to start scraping...")

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            label = f"{sheet_name} Baseball"

            # Collect rows that need processing before iterating so we know
            # the total count for progress output.
            rows_to_process = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                page_url_cell = ws.cell(row=row_idx, column=COL_PAGE_URL)
                download_url_cell = ws.cell(row=row_idx, column=COL_DOWNLOAD_URL)

                page_url = page_url_cell.value
                download_url = download_url_cell.value

                # Only queue rows that have a page URL but no download URL yet.
                if page_url and not download_url:
                    # Try to get a human-readable set name from Column A.
                    set_name = ws.cell(row=row_idx, column=1).value or f"Row {row_idx}"
                    rows_to_process.append((row_idx, str(page_url).strip(), set_name))

            total_in_sheet = len(rows_to_process)

            if total_in_sheet == 0:
                print(f"\n[{label}] All rows already have download URLs — nothing to do.")
                continue

            print(f"\n[{label}] {total_in_sheet} rows to process.")

            for position, (row_idx, page_url, set_name) in enumerate(rows_to_process, start=1):
                prefix = f"[{label}] {position}/{total_in_sheet} — {set_name}"

                try:
                    href = scrape_download_url(page, page_url)

                    if href:
                        ws.cell(row=row_idx, column=COL_DOWNLOAD_URL).value = href
                        # Save immediately — if we crash on row 300, rows 1-299 are safe.
                        wb.save(XLSX_PATH)
                        print(f"{prefix} — DONE ({href})")
                        total_done += 1
                    else:
                        # Page loaded but no download link was present.
                        print(f"{prefix} — FAILED (no download link found on page)")
                        total_failed += 1

                except PlaywrightTimeout:
                    print(f"{prefix} — FAILED (page timed out after {PAGE_LOAD_TIMEOUT / 1000}s)")
                    total_failed += 1

                except Exception as exc:
                    # Catch-all so one bad row never kills the whole run.
                    print(f"{prefix} — FAILED ({type(exc).__name__}: {exc})")
                    total_failed += 1

                # Polite delay between requests.
                time.sleep(REQUEST_DELAY)

        browser_context.close()

    # ── Final summary ──────────────────────────────────────────────────────────
    total_processed = total_done + total_failed
    print("\n" + "─" * 60)
    print("SUMMARY")
    print(f"  Done (URL written):   {total_done}")
    print(f"  Failed (no link):     {total_failed}")
    print(f"  Skipped (had URL):    {total_skipped}")
    print(f"  Total rows touched:   {total_processed}")
    print("─" * 60)

    if total_failed > 0:
        print(f"\n{total_failed} row(s) failed. Rerun the script to retry them —")
        print("skipped rows are rows that already have a URL, so failed rows")
        print("(Column C still empty) will be picked up automatically.")


if __name__ == "__main__":
    main()
